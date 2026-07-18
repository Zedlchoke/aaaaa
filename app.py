import pandas as pd
import unidecode
import re
import os
import threading
from collections import defaultdict, deque
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

COL_TEN_CTY = 'TENDTPN'
COL_MA = 'MADTPN'
COL_DIENGIAI = 'DIENGIAI'

ALIAS_WORDS = {r'\bbetong\b': 'be tong'}
ACTION_VERBS = ['chuyen tien', 'chuyen khoan', 'ck', 'thanh toan', 'tt', 'tra tien', 'ct']
BASE_STOP_WORDS = [r'\bchuyen khoan\b', r'\bck\b', r'\bthanh toan\b', r'\btt\b', r'\brut sec\b', r'\bsec\b', r'\brut tien\b', r'\bnop tien\b', r'\bvao tk\b', r'\bphi dv\b', r'\bphi quan ly\b', r'\bsms banking\b', r'\bhoa don\b', r'\bhd\b', r'\bdot\b', r'\bcuoi\b', r'\btam ung\b', r'\bquyet toan\b', r'\bhdtc\b', r'\bhdkt\b', r'\bbe tong\b', r'\bbetong\b', r'\bly tam\b', r'\bthep\b', r'\bauto\b', r'\bcong ty\b', r'\bcty\b', r'\bctcp\b', r'\bco phan\b', r'\bcp\b', r'\btrach nhiem huu han\b', r'\btnhh\b', r'\bmtv\b', r'\bm t v\b', r'\bmot thanh vien\b', r'\bchi nhanh\b', r'\btap doan\b', r'\blien hop\b', r'\bco so\b', r'\bdoanh nghiep\b', r'\bviet nam\b', r'\bvn\b', r'\bva\b', r'\bgroup\b', r'\bholdings\b', r'\bthuong mai\b', r'\btm\b', r'\bdich vu\b', r'\bdv\b', r'\btmdv\b', r'\btmcp\b', r'\bsan xuat\b', r'\bsx\b', r'\bxuat nhap khau\b', r'\bxnk\b', r'\bdau tu\b', r'\bxay dung\b', r'\bxd\b', r'\bcong nghiep\b', r'\bcn\b', r'\bco khi\b', r'\bkim loai\b', r'\bngu kim\b', r'\bvat lieu\b', r'\bdanh bong\b', r'\bkhuon mau\b', r'\bgia gia cong\b', r'\bkho bai\b', r'\bbao ve\b', r'\bkhoa hoc\b', r'\bcong nghe\b', r'\bmoi truong\b', r'\bmachinery\b', r'\bmetal\b', r'\bnhom hop kim\b', r'\bnhom\b', r'\bhop kim\b', r'\bthiet bi dien\b', r'\bdien\b', r'\bthiet bi\b', r'\bphat trien\b', r'\bky thuat\b', r'\btong hop\b', r'\bquoc te\b', r'\bche tao\b', r'\bvan tai\b', r'\bvat tu\b', r'\bphu lieu\b', r'\bnhua\b', r'\bbao bi\b', r'\btrang tri\b']

ALIAS_PATTERNS = tuple((re.compile(pattern), replacement) for pattern, replacement in ALIAS_WORDS.items())
ACTION_VERBS_PATTERN = '|'.join(ACTION_VERBS)
RE_NON_ALNUM = re.compile(r'[^a-z0-9\s]')
RE_SPACES = re.compile(r'\s+')
RE_EXT_INVALID = re.compile(r'[^a-z0-9\s/\-\+\,]')
RE_MATHANG_CLEAN = re.compile(r'[^A-Z0-9]')
RE_TEMP_STOP_SPLIT = re.compile(r'[,;|\n\r]+')


# ================= TÔ MÀU NGHIỆP VỤ THEO DIỄN GIẢI =================
# Chỉ dùng để tô màu dòng trong file sao kê xuất ra; không tham gia matching,
# không ảnh hưởng thuật toán P1-P9, winner-takes-all hay dữ liệu ghi kết quả.
SPECIAL_HIGHLIGHT_PRIORITY = ("LUONG", "BAO_HIEM", "THUE", "PHI_SMS_NGAN_HANG")
SPECIAL_HIGHLIGHT_FILLS = {
    "LUONG": PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid"),              # đỏ nhạt
    "BAO_HIEM": PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid"),           # xanh lá nhạt
    "THUE": PatternFill(start_color="FFCFE2F3", end_color="FFCFE2F3", fill_type="solid"),               # xanh dương nhạt
    "PHI_SMS_NGAN_HANG": PatternFill(start_color="FFFCE5F1", end_color="FFFCE5F1", fill_type="solid"),  # hồng nhạt
}
SPECIAL_HIGHLIGHT_PATTERNS = {
    # Tránh bắt nhầm "số lượng" bằng (?<!so ) trước từ luong.
    "LUONG": (
        re.compile(r"\btra\s+luong\b"),
        re.compile(r"\bchuyen\s+luong\b"),
        re.compile(r"\bthanh\s+toan\s+luong\b"),
        re.compile(r"\btt\s+luong\b"),
        re.compile(r"\bchi\s+luong\b"),
        re.compile(r"\btien\s+luong\b"),
        re.compile(r"(?<!so )\bluong\b"),
    ),
    "BAO_HIEM": (
        re.compile(r"\bbao\s+hiem\b"),
        re.compile(r"\bbhxh\b"),
        re.compile(r"\bbhyt\b"),
        re.compile(r"\bbhtn\b"),
    ),
    "THUE": (
        re.compile(r"\bthue\b"),
        re.compile(r"\btncn\b"),
        re.compile(r"\btndn\b"),
        re.compile(r"\bgtgt\b"),
    ),
    "PHI_SMS_NGAN_HANG": (
        # SMS được tô hồng dù có hay không có chữ phí, vì người dùng yêu cầu nhận kiểu "sms".
        re.compile(r"\bsms\b"),
        re.compile(r"\btin\s+nhan\s+ngan\s+hang\b"),
        # Với phí ngân hàng, bắt buộc phải có chữ "phi" đứng trước.
        re.compile(r"\bphi\s+(?:sms|tin\s+nhan|ngan\s+hang|bank|banking|internet\s+banking|mobile\s+banking|ibanking|e\s*banking|dv|dich\s+vu|quan\s+ly|chuyen\s+tien|ck|ct)\b"),
        re.compile(r"\bphi\b.{0,40}\bngan\s+hang\b"),
    ),
}


def detect_special_highlight(text):
    """Trả về mã loại tô màu nghiệp vụ, hoặc None nếu dòng không thuộc nhóm nào."""
    text_norm = normalize_basic(text)
    if not text_norm:
        return None
    for highlight_key in SPECIAL_HIGHLIGHT_PRIORITY:
        for pattern in SPECIAL_HIGHLIGHT_PATTERNS[highlight_key]:
            if pattern.search(text_norm):
                return highlight_key
    return None


def apply_row_fill(ws, row_number, fill):
    """Tô màu nguyên dòng đang tồn tại trong worksheet."""
    if fill is None:
        return
    for column_number in range(1, ws.max_column + 1):
        ws.cell(row=row_number, column=column_number).fill = fill
# =====================================================================

RE_WORD_TOKEN = re.compile(r'\b[a-z0-9]+\b')
ACTION_VERBS_REGEX = re.compile(r'\b(?:' + ACTION_VERBS_PATTERN + r')\b')


class AhoMatcher:
    """Multi-pattern matcher; returns the candidate with the smallest original-order key."""

    def __init__(self, keep_all_outputs=False):
        self.keep_all_outputs = keep_all_outputs
        self.children = [{}]
        self.failure = [0]
        self.own_best = [None]
        self.best_output = [None]
        self.all_outputs = [()]
        self._built = False

    def add(self, pattern, key, payload):
        if not pattern:
            return
        state = 0
        for character in pattern:
            next_state = self.children[state].get(character)
            if next_state is None:
                next_state = len(self.children)
                self.children[state][character] = next_state
                self.children.append({})
                self.failure.append(0)
                self.own_best.append(None)
                self.best_output.append(None)
                self.all_outputs.append(())
            state = next_state
        candidate = (key, payload, len(pattern))
        current = self.own_best[state]
        if current is None or key < current[0]:
            self.own_best[state] = candidate

    def build(self):
        queue = deque()
        self.best_output[0] = self.own_best[0]
        self.all_outputs[0] = ((self.own_best[0],) if self.own_best[0] is not None else ())
        for child_state in self.children[0].values():
            self.failure[child_state] = 0
            queue.append(child_state)

        while queue:
            state = queue.popleft()
            fail_state = self.failure[state]
            own = self.own_best[state]
            if self.keep_all_outputs:
                inherited = self.all_outputs[fail_state]
                outputs = ((own,) if own is not None else ()) + inherited
                self.all_outputs[state] = tuple(sorted(outputs, key=lambda item: item[0]))
            else:
                inherited = self.best_output[fail_state]
                if own is None:
                    self.best_output[state] = inherited
                elif inherited is None or own[0] <= inherited[0]:
                    self.best_output[state] = own
                else:
                    self.best_output[state] = inherited

            for character, child_state in self.children[state].items():
                fallback = self.failure[state]
                while fallback and character not in self.children[fallback]:
                    fallback = self.failure[fallback]
                self.failure[child_state] = self.children[fallback].get(character, 0)
                queue.append(child_state)
        self._built = True

    def find_best(self, text, boundary_check=None):
        if not self._built or not text:
            return None
        state = 0
        winner = None
        for end_index, character in enumerate(text):
            while state and character not in self.children[state]:
                state = self.failure[state]
            state = self.children[state].get(character, 0)

            if self.keep_all_outputs:
                for candidate in self.all_outputs[state]:
                    if winner is not None and candidate[0] >= winner[0]:
                        break
                    start_index = end_index - candidate[2] + 1
                    if boundary_check is None or boundary_check(start_index, end_index, candidate):
                        winner = candidate
                        if candidate[0] == 0:
                            return candidate
            else:
                candidate = self.best_output[state]
                if candidate is not None and (winner is None or candidate[0] < winner[0]):
                    winner = candidate
                    # Tuple key (0, 0) and integer key 0 are absolute minima.
                    if candidate[0] == 0 or candidate[0] == (0, 0):
                        return candidate
        return winner


def apply_alias(text):
    for pattern, replacement in ALIAS_PATTERNS:
        text = pattern.sub(replacement, text)
    return text


def normalize_basic(text):
    if pd.isna(text):
        return ""
    normalized = unidecode.unidecode(str(text)).lower().strip()
    normalized = apply_alias(normalized)
    normalized = RE_NON_ALNUM.sub(' ', normalized)
    return RE_SPACES.sub(' ', normalized).strip()


def get_core_name(text, stop_words_list):
    normalized = normalize_basic(text)
    for _ in range(2):
        for stop_word in stop_words_list:
            normalized = re.sub(stop_word, ' ', normalized)
    return RE_SPACES.sub(' ', normalized).strip()


def parse_temporary_owner_names(owner_names_text):
    """Chuẩn hóa và loại trùng tên chủ tài khoản; không lưu ra cấu hình."""
    seen = set()
    result = []
    for raw_name in RE_TEMP_STOP_SPLIT.split(owner_names_text or ""):
        normalized_name = normalize_basic(raw_name)
        if normalized_name and normalized_name not in seen:
            seen.add(normalized_name)
            result.append(normalized_name)
    return result


def build_temporary_owner_variants(owner_names_text):
    """Trả về tên đầy đủ và tên lõi của chủ tài khoản, đã chuẩn hóa và loại trùng."""
    temporary_names = parse_temporary_owner_names(owner_names_text)
    owner_variants = []
    seen_variants = set()
    for name in temporary_names:
        core_name = get_core_name(name, BASE_STOP_WORDS)
        for variant in (name, core_name):
            if variant and variant not in seen_variants:
                seen_variants.add(variant)
                owner_variants.append(variant)
    return temporary_names, owner_variants


def build_temporary_stop_words(owner_names_text):
    """Tạo stop words riêng cho đúng một lần chạy, không sửa BASE_STOP_WORDS."""
    temporary_names, owner_variants = build_temporary_owner_variants(owner_names_text)

    # Đặt tên chủ tài khoản trước stop words nền để tên pháp lý đầy đủ vẫn được
    # loại bỏ trước khi các cụm như "công ty", "TNHH", "MTV" bị xóa riêng lẻ.
    dynamic_stops = [r'\b' + re.escape(name) + r'\b' for name in owner_variants]
    dynamic_stops.extend(BASE_STOP_WORDS)
    return dynamic_stops, temporary_names

def get_bigrams(string):
    s = string.replace(' ', '')
    return [s[i:i+2] for i in range(len(s)-1)] if len(s) > 1 else [s]

def format_excel_sheet(ws, is_doichieu=False):
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="center", wrap_text=True)
    top_alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = body_alignment
    if is_doichieu:
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 55
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 22
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 14
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in [2, 3, 4, 5, 6, 7]:
                cell = row[col_idx - 1]
                cell.alignment = top_alignment
    else:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

def parse_amt_to_float(val):
    if pd.isna(val): return 0.0
    s = str(val).strip().replace(',', '').replace(' ', '')
    if not s: return 0.0
    try: return float(s)
    except:
        m = re.search(r'-?\d+(\.\d+)?', s)
        return float(m.group(0)) if m else 0.0

def load_smart_ktsc(file_path):
    if not file_path or not os.path.exists(file_path): return []
    try:
        with pd.ExcelFile(file_path) as excel_file:
            sheet_name = 'Smart_KTSC_OK' if 'Smart_KTSC_OK' in excel_file.sheet_names else excel_file.sheet_names[0]
            df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
        items = df.to_dict('records')
        clean_items = []
        for row in items:
            clean_items.append({'SO_HD': str(row.get('SO_HD', '')).strip(), 'MATHANG': str(row.get('MATHANG', '')).strip(), 'TTVND': parse_amt_to_float(row.get('TTVND', 0)), 'TTVND_TT': parse_amt_to_float(row.get('TTVND_TT', 0)), COL_MA: str(row.get('MAKH', '')).strip(), COL_TEN_CTY: str(row.get('TENKH', '')).strip()})
        return clean_items
    except Exception as e:
        print(f"Lỗi đọc file: {e}"); return []

# ================= HỆ THỐNG GOM NHÓM & PHÂN LUỒNG =================
def gom_nhom_bang_chung(ket_qua_dict, chuoi_uu_tien):
    p_win_chinh = None
    for p in chuoi_uu_tien:
        if p in ket_qua_dict:
            p_win_chinh = p
            break
            
    if not p_win_chinh:
        return None, "No Match", {}
        
    master_chot = ket_qua_dict[p_win_chinh][2]
    ma_chot = str(master_chot.get(COL_MA, '')).strip()
    ten_chot = str(master_chot.get(COL_TEN_CTY, '')).strip()
    
    # CHỈNH SỬA: Chuẩn hóa không dấu, chữ thường để đối chiếu chéo
    ma_chot_norm = unidecode.unidecode(ma_chot).lower()
    ten_chot_norm = unidecode.unidecode(ten_chot).lower()
    
    cac_p_thang = []
    for p in chuoi_uu_tien:
        if p in ket_qua_dict:
            ma_p = str(ket_qua_dict[p][2].get(COL_MA, '')).strip()
            ten_p = str(ket_qua_dict[p][2].get(COL_TEN_CTY, '')).strip()
            
            # CHỈNH SỬA: Chuẩn hóa kết quả quét được
            ma_p_norm = unidecode.unidecode(ma_p).lower()
            ten_p_norm = unidecode.unidecode(ten_p).lower()
            
            # CHỈNH SỬA: Chỉ so sánh trên nền tảng chữ đã bỏ dấu
            if ma_p_norm == ma_chot_norm and ten_p_norm == ten_chot_norm:
                cac_p_thang.append(p)
                
    ghi_chu = ",".join(cac_p_thang)
    bang_chung = {"ten_quet": [], "hd_quet": [], "hopdong_quet": [], "sotien_quet": []}
    
    for p_win in cac_p_thang:
        pl = ket_qua_dict[p_win][0]
        text = ket_qua_dict[p_win][3]
        if pl in (1, 2, 3, 4, 5, 9) and text: 
            bang_chung["ten_quet"].append(text)
        elif pl == 6 and text: 
            bang_chung["hd_quet"].append(text.replace("SỐ HĐ: ", ""))
        elif pl == 7 and text: 
            bang_chung["hopdong_quet"].append(text.replace("HỢP ĐỒNG: ", ""))
        elif pl == 8 and text: 
            bang_chung["sotien_quet"].append(text.replace("SỐ TIỀN: ", ""))
        
    for k in bang_chung: 
        bang_chung[k] = " | ".join(dict.fromkeys(bang_chung[k]))
        
    return master_chot, ghi_chu, bang_chung
def luong_tuy_chinh(giao_dich, danh_sach_cong_ty, chuoi_uu_tien):
    kq_all = {}
    for p in chuoi_uu_tien:
        kq = chay_thuat_toan(p, giao_dich, danh_sach_cong_ty)
        if kq: kq_all[p] = kq
    return gom_nhom_bang_chung(kq_all, chuoi_uu_tien)

def luong_mac_dinh(giao_dich, danh_sach_cong_ty):
    chuoi_uu_tien = ['P9','P6', 'P7', 'P8', 'P1', 'P2', 'P3', 'P4', 'P5']
    kq_all = {}
    
    for p in chuoi_uu_tien:
        kq = chay_thuat_toan(p, giao_dich, danh_sach_cong_ty)
        if kq: 
            kq_all[p] = kq
            
    # Gom nhóm các bằng chứng từ tất cả các thuật toán ra kết quả
    return gom_nhom_bang_chung(kq_all, chuoi_uu_tien)

global_chay_thuat_toan_func = None
def chay_thuat_toan(p, giao_dich, danh_sach_cong_ty):
    if global_chay_thuat_toan_func:
        return global_chay_thuat_toan_func(p)
    return None
# =================================================================

def process_bank_data(file_saoke, file_master, path_save_doichieu, path_save_saokemoi, user_stop_str, file_muavao=None, file_banra=None, progress_callback=None, enabled_ps=None, custom_order=None):
    """
    Xử lý đối soát giữ nguyên toàn bộ thuật toán P1-P9.

    Các tối ưu chỉ gồm:
    - tiền xử lý/caching dữ liệu bất biến;
    - tránh biên dịch regex và tạo chuỗi lặp lại trong từng vòng master;
    - dùng itertuples thay iterrows;
    - cache kết quả tra cứu P7/P8 và index tên cho P9;
    - stop words do người dùng nhập chỉ tồn tại trong lần chạy hiện tại.
    """
    dynamic_stops, temporary_owner_names = build_temporary_stop_words(user_stop_str)
    _, temporary_owner_variants = build_temporary_owner_variants(user_stop_str)
    compiled_stop_patterns = tuple(re.compile(pattern) for pattern in dynamic_stops)
    compiled_owner_pattern = (
        re.compile('|'.join(r'\b' + re.escape(name) + r'\b' for name in temporary_owner_variants))
        if temporary_owner_variants else None
    )

    def exact_get_core_name(text):
        # Giữ đúng thứ tự và đúng 2 lượt loại stop-word như mã gốc.
        t = normalize_basic(text)
        for _ in range(2):
            for pattern in compiled_stop_patterns:
                t = pattern.sub(' ', t)
        return RE_SPACES.sub(' ', t).strip()

    if progress_callback:
        stop_note = f"; {len(temporary_owner_names)} tên chủ TK tạm thời" if temporary_owner_names else ""
        progress_callback(5, f"Đang đọc file Excel{stop_note}...")

    df_bank = pd.read_excel(file_saoke)
    df_master = pd.read_excel(file_master)
    master_list = df_master.to_dict('records')

    # Tiền xử lý Master một lần. Thứ tự master_list được giữ nguyên tuyệt đối.
    bigram_master_index = defaultdict(list)
    p1_matcher = AhoMatcher(keep_all_outputs=True)
    p4_matcher = AhoMatcher()
    p2_phrase_index = {}
    p3_acronym_index = {}
    for master_index, item in enumerate(master_list):
        item['norm_core'] = exact_get_core_name(item.get(COL_TEN_CTY, ""))
        core = item['norm_core']
        core_words = core.split()
        item['core_words'] = core_words

        if len(core) >= 3:
            p1_pat = r'\b' + r'\s*'.join(map(re.escape, core_words)) + r'\b'
            item['p1_regex'] = re.compile(p1_pat)
            p1_matcher.add(''.join(core_words), master_index, master_index)
        else:
            item['p1_regex'] = None

        # P2: trên chuỗi đã normalize, regex cũ tương đương một cụm từ nguyên vẹn
        # gồm các token liên tiếp. Lưu master đầu tiên cho từng cụm để giữ thứ tự cũ.
        for phrase_start in range(len(core_words)):
            phrase_parts = []
            for phrase_end in range(phrase_start, len(core_words)):
                phrase_parts.append(core_words[phrase_end])
                phrase = ' '.join(phrase_parts)
                if phrase not in p2_phrase_index:
                    p2_phrase_index[phrase] = master_index

        if len(core_words) >= 4:
            acronym = "".join(word[0] for word in core_words)
            item['acronym'] = acronym
            item['p3_regex'] = (
                re.compile(r'\b(?:' + ACTION_VERBS_PATTERN + r')\b.*?\b' + re.escape(acronym) + r'\b')
                if len(acronym) >= 4 else None
            )
            if len(acronym) >= 4 and acronym not in p3_acronym_index:
                p3_acronym_index[acronym] = master_index
        else:
            item['acronym'] = ""
            item['p3_regex'] = None

        # Giữ đúng thứ tự i/j của P4 trong mã gốc.
        p4_chunks = []
        for i in range(len(core_words)):
            for j in range(i + 1, len(core_words) + 1):
                chunk = "".join(core_words[i:j])
                if ((j - i >= 2 and len(chunk) >= 5) or len(chunk) >= 8):
                    p4_chunks.append(chunk)
        item['p4_chunks'] = p4_chunks

        bigram_set = frozenset(get_bigrams(core))
        item['bigram_set'] = bigram_set
        for chunk_order, chunk in enumerate(p4_chunks):
            p4_matcher.add(chunk, (master_index, chunk_order), (master_index, chunk))

        for bigram in bigram_set:
            bigram_master_index[bigram].append(master_index)

    p1_matcher.build()
    p4_matcher.build()

    list_muavao = load_smart_ktsc(file_muavao)
    list_banra = load_smart_ktsc(file_banra)

    def build_hd_index(items):
        result = {}
        for item in items:
            hd = str(item.get('SO_HD', '')).strip().lstrip('0') or '0'
            if hd != '0':
                result.setdefault(hd, []).append(item)
        return result

    hd_index_muavao = build_hd_index(list_muavao)
    hd_index_banra = build_hd_index(list_banra)

    # Tiền xử lý Mua/Bán cho P7/P9, không sửa dữ liệu đầu vào và không đổi thứ tự.
    for item in list_muavao + list_banra:
        ten_cty = str(item.get(COL_TEN_CTY, '')).strip()
        item['tenkh_norm'] = unidecode.unidecode(ten_cty).lower() if ten_cty else ""
        mathang_upper = str(item.get('MATHANG', '')).upper()
        item['mathang_upper'] = mathang_upper
        item['mathang_clean'] = RE_MATHANG_CLEAN.sub('', mathang_upper)

    for master_item in master_list:
        ten_cty_master = str(master_item.get(COL_TEN_CTY, '')).strip()
        master_item['tenkh_norm'] = unidecode.unidecode(ten_cty_master).lower() if ten_cty_master else ""

    # Các tập hợp đích được tạo một lần thay vì ghép lại trên từng dòng.
    combined_target_list = list_banra + list_muavao
    combined_hd_index = {**hd_index_banra, **hd_index_muavao}

    target_lists = {
        'mua': list_muavao,
        'ban': list_banra,
        'all': combined_target_list,
    }
    target_hd_indexes = {
        'mua': hd_index_muavao,
        'ban': hd_index_banra,
        'all': combined_hd_index,
    }

    def build_first_name_index(items):
        index = {}
        for item in items:
            normalized_name = item.get('tenkh_norm', '')
            if normalized_name and normalized_name not in index:
                index[normalized_name] = item
        return index

    target_name_indexes = {kind: build_first_name_index(items) for kind, items in target_lists.items()}
    master_name_index = build_first_name_index(master_list)

    hd_pattern_comp = re.compile(r'\b(?:hoa don|hd)\s*(?:so\s*)?((?:\d+(?:\s*(?:,|\+|va\b|-)\s*)*)+)')
    p7_hopdong_pattern = re.compile(r'\bhop dong\s*(?:so\s*)?([a-z0-9/\-]+)\b')
    p7_hopdong_spaced_pattern = re.compile(r'\bhop dong\s*(?:so\s*)?\d+\s+([a-z0-9/\-]+)\b')
    p7_hd_pattern = re.compile(r'\bhd\s*(?:so\s*)?([a-z0-9/\-]+)\b')

    if enabled_ps is None:
        enabled_ps = {"P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "P9"}
    else:
        enabled_ps = {p.upper() for p in enabled_ps}

    if custom_order:
        try:
            order_list = [p.strip().upper() for p in str(custom_order).split(',') if p.strip()]
        except Exception:
            order_list = None
    else:
        order_list = None

    total_rows = len(df_bank)
    h_d_bank = {str(column).strip().upper(): column for column in df_bank.columns}
    col_tkno_name = h_d_bank.get('TKNO')
    col_tkco_name = h_d_bank.get('TKCO')
    col_tenkh_name_in_df = h_d_bank.get('TENKH')

    if COL_DIENGIAI not in df_bank.columns:
        raise ValueError(f"File sao kê thiếu cột bắt buộc: {COL_DIENGIAI}")

    if progress_callback:
        progress_callback(10, "Đang tiền xử lý dữ liệu hàng loạt...")

    # Giữ nguyên cách combined regex của mã cũ cho DIENGIAI_CLEANED.
    compiled_stop_pattern = re.compile('|'.join(dynamic_stops))

    def fast_get_core_name(text_norm):
        t = compiled_stop_pattern.sub(' ', text_norm)
        t = compiled_stop_pattern.sub(' ', t)
        return RE_SPACES.sub(' ', t).strip()

    def remove_temporary_owner(text_norm):
        # Chỉ loại tên chủ tài khoản; không loại BASE_STOP_WORDS ở luồng P1/P3/P4,
        # vì các thuật toán này vốn dùng DIENGIAI_NORM của hệ thống cũ.
        if compiled_owner_pattern is None:
            return text_norm
        cleaned = compiled_owner_pattern.sub(' ', text_norm)
        cleaned = compiled_owner_pattern.sub(' ', cleaned)
        return RE_SPACES.sub(' ', cleaned).strip()

    def fast_text_ext(text):
        t = unidecode.unidecode(str(text)).lower()
        t = RE_EXT_INVALID.sub(' ', t)
        return RE_SPACES.sub(' ', t).strip()

    # Dùng list comprehension thay Series.apply để giảm overhead Python/Pandas.
    diengiai_goc_values = df_bank[COL_DIENGIAI].astype(str).tolist()
    diengiai_norm_values = [normalize_basic(value) for value in diengiai_goc_values]
    diengiai_owner_filtered_values = [remove_temporary_owner(value) for value in diengiai_norm_values]
    diengiai_cleaned_values = [fast_get_core_name(value) for value in diengiai_norm_values]
    diengiai_nospace_values = [value.replace(' ', '') for value in diengiai_owner_filtered_values]
    diengiai_ext_values = [fast_text_ext(value) for value in diengiai_goc_values]

    columns = list(df_bank.columns)
    exact_column_positions = {column: index for index, column in enumerate(columns)}
    tkno_pos = exact_column_positions.get(col_tkno_name) if col_tkno_name is not None else None
    tkco_pos = exact_column_positions.get(col_tkco_name) if col_tkco_name is not None else None
    tenkh_pos = exact_column_positions.get(col_tenkh_name_in_df) if col_tenkh_name_in_df is not None else None
    ttvnd_pos = exact_column_positions.get('TTVND')
    ttvnd_tt_pos = exact_column_positions.get('TTVND_TT')

    all_matches = []
    p7_cache = {}
    p8_cache = {}
    progress_step = max(20, total_rows // 500 if total_rows else 20)

    for idx, values in enumerate(df_bank.itertuples(index=False, name=None)):
        if progress_callback and idx % progress_step == 0:
            progress_callback(15 + (idx / max(total_rows, 1)) * 65, f"Đối soát dòng {idx + 1}/{total_rows}...")

        thu_tu_dong = idx + 2
        diengiai_goc = diengiai_goc_values[idx]

        if not diengiai_goc or diengiai_goc.strip() == "" or diengiai_goc.lower().strip() == "nan":
            all_matches.append({"THỨ TỰ DÒNG GỐC": thu_tu_dong, "DIỄN GIẢI GỐC": "", "TÊN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "", "HÓA ĐƠN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "", "HỢP ĐỒNG QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "", "SỐ TIỀN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "", "TÊN MATCH ĐƯỢC TRẠNG FILE ĐỐI TƯỢNG PHÁP NHÂN": "", "MÃ ĐỐI TƯỢNG PHÁP NHÂN": "", "CÁCH MATCH": "", "SCORE_NUM": 0})
            continue

        diengiai_norm = diengiai_owner_filtered_values[idx]
        diengiai_cleaned = diengiai_cleaned_values[idx]
        diengiai_nospace = diengiai_nospace_values[idx]
        text_ext_precalculated = diengiai_ext_values[idx]

        amt_val = parse_amt_to_float(values[ttvnd_pos]) if ttvnd_pos is not None else 0.0
        if amt_val == 0:
            amt_val = parse_amt_to_float(values[ttvnd_tt_pos]) if ttvnd_tt_pos is not None else 0.0

        tkco_val = str(values[tkco_pos]).strip() if tkco_pos is not None else ""
        tkno_val = str(values[tkno_pos]).strip() if tkno_pos is not None else ""
        is_mua_vao = tkco_val.startswith('112')
        is_ban_ra = tkno_val.startswith('112')

        if is_mua_vao and list_muavao:
            target_kind = 'mua'
        elif is_ban_ra and list_banra:
            target_kind = 'ban'
        else:
            target_kind = 'all'

        current_target_list = target_lists[target_kind]
        current_hd_index = target_hd_indexes[target_kind]
        current_name_index = target_name_indexes[target_kind]
        matches_for_row = []

        # P2 dùng cùng một regex cho toàn bộ master của một dòng; chỉ biên dịch một lần.
        p2_regex = None
        if len(diengiai_cleaned) >= 4:
            p2_regex = re.compile(r'\b' + r'\s+'.join(map(re.escape, diengiai_cleaned.split())) + r'\b')

        def try_p1():
            if "P1" not in enabled_ps:
                return

            # Aho chỉ tạo tập ứng viên. Regex gốc xác nhận ứng viên để giữ chính xác
            # word-boundary và quy tắc \s* của P1, kể cả chuỗi số có khoảng trắng.
            def p1_exact_check(start_index, end_index, candidate):
                master_item = master_list[candidate[1]]
                p1_regex = master_item.get('p1_regex')
                return p1_regex is not None and p1_regex.search(diengiai_norm) is not None

            candidate = p1_matcher.find_best(diengiai_nospace, p1_exact_check)
            if candidate is not None:
                master_item = master_list[candidate[1]]
                core = master_item['norm_core']
                matches_for_row.append((1, len(core), master_item, core.upper()))

        def try_p2():
            if "P2" not in enabled_ps or len(diengiai_cleaned) < 4:
                return
            master_index = p2_phrase_index.get(diengiai_cleaned)
            if master_index is not None:
                master_item = master_list[master_index]
                core = master_item['norm_core']
                matches_for_row.append((2, -len(core), master_item, diengiai_cleaned.upper()))

        def try_p3():
            if "P3" not in enabled_ps:
                return
            action_match = ACTION_VERBS_REGEX.search(diengiai_norm)
            if action_match is None:
                return
            best_master_index = None
            for token in RE_WORD_TOKEN.findall(diengiai_norm[action_match.end():]):
                master_index = p3_acronym_index.get(token)
                if master_index is not None and (best_master_index is None or master_index < best_master_index):
                    best_master_index = master_index
            if best_master_index is not None:
                master_item = master_list[best_master_index]
                acronym = master_item['acronym']
                matches_for_row.append((3, len(acronym), master_item, acronym.upper()))

        def try_p4():
            if "P4" not in enabled_ps:
                return
            candidate = p4_matcher.find_best(diengiai_nospace)
            if candidate is not None:
                master_index, chunk = candidate[1]
                master_item = master_list[master_index]
                matches_for_row.append((4, len(chunk), master_item, chunk.upper()))

        def try_p5():
            if "P5" not in enabled_ps or len(diengiai_cleaned) < 5:
                return
            set1 = set(get_bigrams(diengiai_cleaned))
            len_set1 = len(set1)
            if len_set1 == 0:
                return

            # Đếm giao bigram bằng posting list. Công thức và thứ tự tie vẫn nguyên.
            intersection_counts = defaultdict(int)
            for bigram in set1:
                for master_index in bigram_master_index.get(bigram, ()):
                    intersection_counts[master_index] += 1

            best_sim = 0.0
            best_master = None
            for master_index in sorted(intersection_counts):
                master_item = master_list[master_index]
                inter_len = intersection_counts[master_index]
                set2 = master_item['bigram_set']
                sim = 2.0 * inter_len / (len_set1 + len(set2))
                if sim > best_sim:
                    best_sim = sim
                    best_master = master_item

            if best_sim >= 0.85 and best_master is not None:
                matches_for_row.append((5, int(best_sim * 100), best_master, f"TÊN AI: {best_master['norm_core'].upper()}"))

        def get_text_ext():
            return text_ext_precalculated

        def try_p6():
            if "P6" not in enabled_ps or not current_target_list or amt_val < 5000000:
                return
            raw_hd_groups = hd_pattern_comp.findall(get_text_ext())
            for hd_group in raw_hd_groups:
                nums = re.findall(r'\d+', hd_group)
                if not nums:
                    continue
                candidate_sets = []
                best_item_map = {}
                for num in set(nums):
                    num_clean = num.lstrip('0') or '0'
                    if num_clean == '0':
                        continue
                    cands = current_hd_index.get(num_clean, [])
                    if cands:
                        mas_set = set()
                        for item in cands:
                            ma = item.get(COL_MA)
                            if ma:
                                mas_set.add(ma)
                                if ma not in best_item_map:
                                    best_item_map[ma] = item
                            elif item.get(COL_TEN_CTY):
                                ten = item.get(COL_TEN_CTY)
                                mas_set.add(ten)
                                if ten not in best_item_map:
                                    best_item_map[ten] = item
                        if mas_set:
                            candidate_sets.append(mas_set)
                if candidate_sets:
                    try:
                        common = set.intersection(*candidate_sets)
                        if len(common) == 1:
                            matched_key = list(common)[0]
                            matches_for_row.append((6, len(nums), best_item_map[matched_key], f"SỐ HĐ: {', '.join(nums)}"))
                            return
                    except Exception:
                        pass

        def find_p7_item(contract):
            cache_key = (target_kind, contract)
            if cache_key in p7_cache:
                return p7_cache[cache_key]

            c_clean = RE_MATHANG_CLEAN.sub('', contract.upper())
            parts = re.split(r'[/_.-]', contract.upper())
            valid_parts = [part for part in parts if len(part) >= 5]
            matched_item = None
            for item in current_target_list:
                mathang_upper = item.get('mathang_upper', '')
                mathang_clean = item.get('mathang_clean', '')
                matched = False
                if len(c_clean) >= 5 and c_clean in mathang_clean:
                    matched = True
                elif any(valid_part in mathang_upper for valid_part in valid_parts):
                    matched = True
                if matched:
                    matched_item = item
                    break

            p7_cache[cache_key] = matched_item
            return matched_item

        def try_p7():
            if "P7" not in enabled_ps or not current_target_list or amt_val < 5000000:
                return
            text_ext = get_text_ext()
            contracts_found = []
            explicit_contracts = p7_hopdong_pattern.findall(text_ext)
            contracts_found.extend([contract for contract in explicit_contracts if len(contract) >= 4])
            explicit_contracts_spaced = p7_hopdong_spaced_pattern.findall(text_ext)
            contracts_found.extend([contract for contract in explicit_contracts_spaced if '/' in contract or '-' in contract])
            hd_matches = p7_hd_pattern.findall(text_ext)
            for value in hd_matches:
                if ('/' in value or '-' in value) and len(value) >= 4:
                    contracts_found.append(value)

            # Giữ nguyên set() và thứ tự duyệt như mã gốc.
            for contract in set(contracts_found):
                matched_item = find_p7_item(contract)
                if matched_item is not None:
                    matches_for_row.append((7, len(contract), matched_item, f"HỢP ĐỒNG: {contract}"))
                    return

        def find_p8_item():
            cache_key = (target_kind, amt_val)
            if cache_key in p8_cache:
                return p8_cache[cache_key]

            matched_companies = []
            best_item = None
            for item in current_target_list:
                ttvnd = item.get('TTVND', 0.0)
                ttvnd_tt = item.get('TTVND_TT', 0.0)
                if abs(ttvnd - amt_val) < 1.0 or abs(ttvnd_tt - amt_val) < 1.0:
                    key = item.get(COL_MA) if item.get(COL_MA) else item.get(COL_TEN_CTY)
                    matched_companies.append(key)
                    best_item = item

            result = best_item if len(set(matched_companies)) == 1 else None
            p8_cache[cache_key] = result
            return result

        def try_p8():
            if "P8" not in enabled_ps or not current_target_list or amt_val < 5000000:
                return
            best_item = find_p8_item()
            if best_item is not None:
                matches_for_row.append((8, 0, best_item, f"SỐ TIỀN: {amt_val:,.0f}"))

        def try_p9():
            if "P9" not in enabled_ps:
                return
            tenkh_goc = str(values[tenkh_pos]) if tenkh_pos is not None else ""
            if not tenkh_goc or tenkh_goc.strip().lower() in ["", "nan"]:
                return
            tenkh_goc_norm = unidecode.unidecode(tenkh_goc).lower().strip()

            target_item = current_name_index.get(tenkh_goc_norm)
            if target_item is not None and tenkh_goc_norm != "":
                matches_for_row.append((9, 100, target_item, "P9 - TRÙNG TÊN KH GỐC (MUA/BÁN)"))
                return

            master_item = master_name_index.get(tenkh_goc_norm)
            if master_item is not None and tenkh_goc_norm != "":
                matches_for_row.append((9, 100, master_item, "P9 - TRÙNG TÊN KH GỐC (MASTER)"))

        def row_chay_thuat_toan(p):
            nonlocal matches_for_row
            matches_for_row = []
            if p == "P1":
                try_p1()
            elif p == "P2":
                try_p2()
            elif p == "P3":
                try_p3()
            elif p == "P4":
                try_p4()
            elif p == "P5":
                try_p5()
            elif p == "P6":
                try_p6()
            elif p == "P7":
                try_p7()
            elif p == "P8":
                try_p8()
            elif p == "P9":
                try_p9()

            if matches_for_row:
                matches_for_row.sort(key=lambda match: (match[0], -match[1]))
                return matches_for_row[0]
            return None

        global global_chay_thuat_toan_func
        global_chay_thuat_toan_func = row_chay_thuat_toan

        if order_list:
            master_chot, ghi_chu_chot, bc = luong_tuy_chinh(None, None, order_list)
        else:
            master_chot, ghi_chu_chot, bc = luong_mac_dinh(None, None)

        if master_chot:
            all_matches.append({
                "THỨ TỰ DÒNG GỐC": thu_tu_dong,
                "DIỄN GIẢI GỐC": diengiai_goc,
                "TÊN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": bc["ten_quet"],
                "HÓA ĐƠN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": bc["hd_quet"],
                "HỢP ĐỒNG QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": bc["hopdong_quet"],
                "SỐ TIỀN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": bc["sotien_quet"],
                "TÊN MATCH ĐƯỢC TRẠNG FILE ĐỐI TƯỢNG PHÁP NHÂN": master_chot.get(COL_TEN_CTY, ""),
                "MÃ ĐỐI TƯỢNG PHÁP NHÂN": master_chot.get(COL_MA, ""),
                "CÁCH MATCH": f"{ghi_chu_chot}",
                "SCORE_NUM": 100,
            })
        else:
            all_matches.append({
                "THỨ TỰ DÒNG GỐC": thu_tu_dong,
                "DIỄN GIẢI GỐC": diengiai_goc,
                "TÊN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "",
                "HÓA ĐƠN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "",
                "HỢP ĐỒNG QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "",
                "SỐ TIỀN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI": "",
                "TÊN MATCH ĐƯỢC TRẠNG FILE ĐỐI TƯỢNG PHÁP NHÂN": "",
                "MÃ ĐỐI TƯỢNG PHÁP NHÂN": "",
                "CÁCH MATCH": "No Match",
                "SCORE_NUM": 0,
            })

    # Không giữ closure của dòng cuối cùng và toàn bộ dữ liệu lớn sau khi match xong.
    global_chay_thuat_toan_func = None

    # Mã gốc sử dụng matches_dict bên dưới nhưng chưa khởi tạo.
    # Đây là khôi phục biến ánh xạ đã được phần cập nhật sao kê mong đợi.
    matches_dict = {match["THỨ TỰ DÒNG GỐC"]: match for match in all_matches}

    wb_doichieu = Workbook()
    ws = wb_doichieu.active
    ws.title = "Ket Qua Match"
    headers = ["THỨ TỰ DÒNG GỐC", "DIỄN GIẢI GỐC", "TÊN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI", "HÓA ĐƠN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI", "HỢP ĐỒNG QUÉT ĐƯỢC TRẠNG DIỄN GIẢI", "SỐ TIỀN QUÉT ĐƯỢC TRẠNG DIỄN GIẢI", "TÊN MATCH ĐƯỢC TRẠNG FILE ĐỐI TƯỢNG PHÁP NHÂN", "MÃ ĐỐI TƯỢNG PHÁP NHÂN", "CÁCH MATCH"]
    ws.append(headers)
    for match in all_matches:
        ws.append([match.get(header, "") for header in headers])
    format_excel_sheet(ws, is_doichieu=True)

    ws_legend = wb_doichieu.create_sheet("Giai Thich P1-P9")
    legend_data = [
        ["CÁCH MATCH", "Ý NGHĨA", "MÔ TẢ CHI TIẾT"],
        ["P1", "Khớp tên nguyên vẹn", "Tìm chính xác tên công ty theo từ khóa (word boundary) trong diễn giải"],
        ["P2", "Khớp ngược", "Tên công ty nằm trong diễn giải (ngược lại)"],
        ["P3", "Khớp viết tắt (acronym)", "Dùng chữ cái đầu của từ trong tên công ty kết hợp động từ (chuyển, tt, ck...)"],
        ["P4", "Khớp cụm từ (chunk)", "Tìm cụm từ dài (từ 2 từ trở lên) trong tên công ty xuất hiện trong diễn giải"],
        ["P5", "Fuzzy AI (bigram)", "So sánh độ tương đồng tên bằng thuật toán AI (bigram similarity ≥ 85%)"],
        ["P6", "Hóa đơn (Mua/Bán)", "Tìm số HÓA ĐƠN trong diễn giải → đối chiếu với file Mua VÀO hoặc BÁN RA"],
        ["P7", "Hợp đồng / MATHANG", "Tìm số HỢP ĐỒNG hoặc tên hàng trong diễn giải → đối chiếu với file Mua VÀO hoặc BÁN RA"],
        ["P8", "Số tiền duy nhất", "Khớp duy nhất theo số tiền khi dòng ≥ 5 triệu và có file Mua/Bán"],
        ["P9", "Khớp Tên KH Gốc", "Dùng TENKH sẵn có trên file sao kê gốc để đối chiếu chéo (không dấu) với file Mua/Bán"],
    ]
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row_index, row_data in enumerate(legend_data, 1):
        for column_index, value in enumerate(row_data, 1):
            cell = ws_legend.cell(row=row_index, column=column_index, value=value)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_index == 1:
                cell.font = Font(bold=True)
    ws_legend.column_dimensions['A'].width = 14
    ws_legend.column_dimensions['B'].width = 25
    ws_legend.column_dimensions['C'].width = 80
    wb_doichieu.save(path_save_doichieu)
    wb_doichieu.close()

    wb_saoke = load_workbook(file_saoke)
    ws = wb_saoke.active
    h_d = {str(cell.value).strip().upper(): cell.column for cell in ws[1] if cell.value}
    col_tkno = h_d.get('TKNO', 5)
    col_tkco = h_d.get('TKCO', 7)
    col_madtpnno = h_d.get('MADTPNNO', 6)
    col_madtpnco = h_d.get('MADTPNCO', 8)
    col_tenkh = h_d.get('TENKH', 11)
    col_ghichu = h_d.get('GHICHU')
    col_diengiai = h_d.get(COL_DIENGIAI)

    def is_valid_cell(value):
        return pd.notna(value) and str(value).strip() != "" and str(value).lower().strip() != "nan"

    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    for row_number in range(2, ws.max_row + 1):
        diengiai_value = ws.cell(row=row_number, column=col_diengiai).value if col_diengiai else ""
        highlight_key = detect_special_highlight(diengiai_value)
        special_fill = SPECIAL_HIGHLIGHT_FILLS.get(highlight_key) if highlight_key else None

        match = matches_dict.get(row_number)
        if match is not None:
            ma_dtpn = str(match.get("MÃ ĐỐI TƯỢNG PHÁP NHÂN", "")).strip()
            if ma_dtpn.lower() == "nan":
                ma_dtpn = ""

            tkco_cell_value = ws.cell(row=row_number, column=col_tkco).value
            tkno_cell_value = ws.cell(row=row_number, column=col_tkno).value
            tkco_val = str(tkco_cell_value).strip() if is_valid_cell(tkco_cell_value) else ""
            tkno_val = str(tkno_cell_value).strip() if is_valid_cell(tkno_cell_value) else ""

            matched_name = str(match.get("TÊN MATCH ĐƯỢC TRẠNG FILE ĐỐI TƯỢNG PHÁP NHÂN", "")).strip()
            if matched_name.lower() == "nan":
                matched_name = ""
            match_method = str(match.get("CÁCH MATCH", "")).strip()
            if match_method.lower() == "nan":
                match_method = ""

            # QUAN TRỌNG: Nếu dòng không match được thì giữ nguyên dữ liệu sao kê gốc.
            # Bản cũ vẫn đưa dòng No Match vào matches_dict, rồi ghi chuỗi rỗng vào
            # MADTPN/TENKH/GHICHU, làm mất TENKH đã có sẵn trước khi xử lý.
            # Từ đây: No Match chỉ được tô màu, tuyệt đối không xóa/ghi đè ô cũ.
            is_successful_match = bool(ma_dtpn or matched_name) and match_method != "No Match"

            if is_successful_match:
                # Không bao giờ ghi đè bằng chuỗi rỗng. Nếu một phần kết quả thiếu,
                # giữ nguyên giá trị cũ của ô đó để tránh mất dữ liệu thủ công/file gốc.
                if ma_dtpn:
                    if tkco_val == "1121" or (tkco_val != "" and tkno_val == ""):
                        ws.cell(row=row_number, column=col_madtpnno).value = ma_dtpn
                    elif tkno_val == "1121" or (tkno_val != "" and tkco_val == ""):
                        ws.cell(row=row_number, column=col_madtpnco).value = ma_dtpn
                    else:
                        ws.cell(row=row_number, column=col_madtpnno).value = ma_dtpn

                if matched_name:
                    ws.cell(row=row_number, column=col_tenkh).value = matched_name
                if col_ghichu and match_method:
                    ws.cell(row=row_number, column=col_ghichu).value = match_method

                # Các dòng nghiệp vụ đặc biệt luôn được tô màu, dù đã match thành công.
                apply_row_fill(ws, row_number, special_fill)
            else:
                # Nếu vừa No Match vừa thuộc nhóm nghiệp vụ đặc biệt, màu nghiệp vụ ưu tiên hơn màu vàng.
                apply_row_fill(ws, row_number, special_fill or yellow_fill)
        else:
            apply_row_fill(ws, row_number, special_fill or yellow_fill)

    if progress_callback:
        progress_callback(98, "Đang lưu file...")
    wb_saoke.save(path_save_saokemoi)
    wb_saoke.close()
    if progress_callback:
        progress_callback(100, "Hoàn tất thành công!")

class AppGomNghiepVu:
    def __init__(self, root):
        self.root = root
        self.root.title("Hệ Thống Xử Lý Kế Toán & Ngân Hàng Toàn Diện ETAX Hồ Chí Minh")
        self.root.geometry("780x670")
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab1, text="  1. Đối Soát Ngân Hàng (Excel)  ")
        self.notebook.add(self.tab2, text="  2. Gom Ảnh Hàng Loạt Thành PDF Chứng Từ  ")
        
        self.setup_tab1_interface()
        self.setup_tab2_interface()

    def setup_tab1_interface(self):
        self.file_saoke_path = self.file_master_path = self.file_muavao_path = self.file_banra_path = None
        tk.Label(self.tab1, text="PHẦN MỀM ĐỐI CHIẾU MÃ PHÁP NHÂN TRÊN EXCEL", font=("Arial", 14, "bold"), fg="#4F81BD").pack(pady=10)
        
        f_files = tk.Frame(self.tab1)
        f_files.pack(fill="x", padx=20, pady=3)
        tk.Button(f_files, text="1. Chọn File Sao Kê (Excel)", command=lambda: self.chon_file("sk1"), width=25).grid(row=0, column=0, pady=3, padx=5)
        self.lbl_saoke = tk.Label(f_files, text="Chưa chọn file (Bắt buộc)...", fg="gray")
        self.lbl_saoke.grid(row=0, column=1, sticky="w")
        tk.Button(f_files, text="2. Chọn Master MADTPN", command=lambda: self.chon_file("mst"), width=25).grid(row=1, column=0, pady=3, padx=5)
        self.lbl_master = tk.Label(f_files, text="Chưa chọn file (Bắt buộc)...", fg="gray")
        self.lbl_master.grid(row=1, column=1, sticky="w")
        tk.Button(f_files, text="3. Chọn File MUA VÀO Năm", command=lambda: self.chon_file("muavao"), width=25).grid(row=2, column=0, pady=3, padx=5)
        self.lbl_muavao = tk.Label(f_files, text="Chưa chọn file (Không bắt buộc)...", fg="gray")
        self.lbl_muavao.grid(row=2, column=1, sticky="w")
        tk.Button(f_files, text="4. Chọn File BÁN RA Năm", command=lambda: self.chon_file("banra"), width=25).grid(row=3, column=0, pady=3, padx=5)
        self.lbl_banra = tk.Label(f_files, text="Chưa chọn file (Không bắt buộc)...", fg="gray")
        self.lbl_banra.grid(row=3, column=1, sticky="w")
        
        f_match = tk.LabelFrame(self.tab1, text=" Tùy chọn kiểu match (P1-P9) - Tick để kích hoạt ", padx=10, pady=5)
        f_match.pack(fill="x", padx=20, pady=5)
        
        self.p_vars = {}
        p_labels = {
            "P1": "P1 - Khớp tên nguyên vẹn",
            "P2": "P2 - Khớp ngược",
            "P3": "P3 - Khớp viết tắt (acronym)",
            "P4": "P4 - Khớp cụm từ (chunk)",
            "P5": "P5 - Fuzzy AI",
            "P6": "P6 - Hóa đơn gộp",
            "P7": "P7 - Hợp đồng / MATHANG",
            "P8": "P8 - Số tiền duy nhất",
            "P9": "P9 - Dùng TENKH gốc trên sao kê"
        }
        for i, (p, label) in enumerate(p_labels.items()):
            row = i // 2
            col = i % 2
            var = tk.BooleanVar(value=True)
            self.p_vars[p] = var
            cb = ttk.Checkbutton(f_match, text=label, variable=var)
            cb.grid(row=row, column=col, sticky="w", padx=5, pady=2)
            
        f_order = tk.Frame(f_match)
        f_order.grid(row=5, column=0, columnspan=2, sticky="w", pady=5)
        tk.Label(f_order, text="Thứ tự tùy chỉnh (trống):", font=("Arial", 9)).pack(side="left")
        self.custom_order_entry = tk.Entry(f_order, width=35, font=("Arial", 9))
        self.custom_order_entry.pack(side="left", padx=5)
        
        f_stop = tk.Frame(self.tab1)
        f_stop.pack(fill="x", padx=20, pady=5)
        tk.Label(f_stop, text="Tên chủ tài khoản cần loại trừ tạm thời (không lưu):", font=("Arial", 10, "bold")).pack(anchor="w")
        self.entry_stop_words = tk.Entry(f_stop, width=70, font=("Arial", 11))
        self.entry_stop_words.pack(anchor="w", ipady=3)
        tk.Label(f_stop, text="Nhập nhiều tên bằng dấu phẩy, chấm phẩy hoặc |. Tên chỉ có hiệu lực trong lần chạy này.", font=("Arial", 8), fg="gray").pack(anchor="w")
        
        self.btn_run_t1 = tk.Button(self.tab1, text="BẤT ĐẦU ĐỐI SOÁT EXCEL", bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), command=self.t1_chay, height=2)
        self.btn_run_t1.pack(fill="x", padx=25, pady=10)
        self.progress_var_t1 = tk.DoubleVar()
        self.progressbar_t1 = ttk.Progressbar(self.tab1, variable=self.progress_var_t1, maximum=100)
        self.progressbar_t1.pack(fill="x", padx=25, pady=5)
        self.lbl_status_t1 = tk.Label(self.tab1, text="Sẵn sàng...", font=("Arial", 9, "italic"))
        self.lbl_status_t1.pack()

    def t1_chay(self):
        if not self.file_saoke_path or not self.file_master_path: 
            return messagebox.showerror("Lỗi", "Vui lòng chọn đủ File Sao kê và File data MADTPN!")
        
        p_dc = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="KetQua_DoiChieu.xlsx")
        if not p_dc: return
        p_sk = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="SaoKe_DaCapNhat.xlsx")
        if not p_sk: return
        
        user_stops = self.entry_stop_words.get()
        enabled = [p for p, var in self.p_vars.items() if var.get()]
        custom_order = self.custom_order_entry.get().strip() or None
        
        self.btn_run_t1.config(state="disabled")
        self.progress_var_t1.set(0)
        
        def task():
            try:
                process_bank_data(
                    file_saoke=self.file_saoke_path, 
                    file_master=self.file_master_path, 
                    path_save_doichieu=p_dc, 
                    path_save_saokemoi=p_sk, 
                    user_stop_str=user_stops, 
                    file_muavao=self.file_muavao_path, 
                    file_banra=self.file_banra_path, 
                    progress_callback=self.t1_update_progress, 
                    enabled_ps=enabled, 
                    custom_order=custom_order
                )
                self.root.after(0, lambda: messagebox.showinfo("Xong", "Đã đối soát thành công!"))
            except Exception as e:
                self.root.after(0, lambda e=e: messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{str(e)}"))
            finally:
                self.root.after(0, lambda: self.btn_run_t1.config(state="normal"))
                
        threading.Thread(target=task, daemon=True).start()

    def setup_tab2_interface(self):
        self.selected_images = []
        tk.Label(self.tab2, text="ỨNG DỤNG GOM HÌNH ẢNH THÀNH PDF CHỨNG TỪ", font=("Arial", 13, "bold"), fg="#2E7D32").pack(pady=15)
        f_btns = tk.Frame(self.tab2)
        f_btns.pack()
        tk.Button(f_btns, text="+ Chọn Thêm Ảnh", command=self.t3_chon_anh, width=32).grid(row=0, column=0, padx=5)
        tk.Button(f_btns, text="🗑 Xóa Toàn Bộ", command=self.t3_xoa_anh, width=25).grid(row=0, column=1, padx=5)
        self.txt_img_list = tk.Text(self.tab2, height=10, width=65, state="disabled", bg="#F5F5F5")
        self.txt_img_list.pack(pady=10)
        self.lbl_so_anh = tk.Label(self.tab2, text="Chưa chọn ảnh.", fg="gray", font=("Arial", 9))
        self.lbl_so_anh.pack()
        self.btn_run_t3 = tk.Button(self.tab2, text="XUẤT RA PDF CHỨNG TỪ", bg="#2E7D32", fg="white", font=("Arial", 11, "bold"), command=self.t3_chay, state="disabled", height=2)
        self.btn_run_t3.pack(fill="x", padx=25, pady=10)

    def t3_chon_anh(self):
        files = filedialog.askopenfilenames(filetypes=[("Image files", "*.png *.jpg *.jpeg")])
        if files: self.selected_images.extend(files); self.t3_cap_nhat_ui()

    def t3_xoa_anh(self):
        self.selected_images = []
        self.t3_cap_nhat_ui()

    def t3_cap_nhat_ui(self):
        self.txt_img_list.config(state="normal")
        self.txt_img_list.delete("1.0", tk.END)
        for i, f_path in enumerate(self.selected_images):
            self.txt_img_list.insert(tk.END, f"{i+1}. {os.path.basename(f_path)}\n")
        self.txt_img_list.config(state="disabled")
        self.btn_run_t3.config(state="normal" if self.selected_images else "disabled")
        self.lbl_so_anh.config(text=f"Đã chọn {len(self.selected_images)} ảnh")

    def t3_chay(self):
        p_pdf = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile="TaiLieu_ChungTu_ETAX.pdf")
        if p_pdf:
            self.btn_run_t3.config(state="disabled", text="ĐANG GHÉP VÀ LƯU PDF...")
            self.root.update()
            imgs = []
            try:
                imgs = [Image.open(p).convert('RGB') for p in self.selected_images]
                if imgs:
                    if len(imgs) == 1: imgs[0].save(p_pdf, "PDF", resolution=100.0)
                    else: imgs[0].save(p_pdf, save_all=True, append_images=imgs[1:], resolution=100.0)
                    messagebox.showinfo("Thành công", f"Đã tạo PDF thành công!\n\n{p_pdf}")
                    self.t3_xoa_anh()
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))
            finally:
                for img in imgs:
                    img.close()
                self.btn_run_t3.config(state="normal", text="XUẤT RA PDF CHỨNG TỪ")

    def t1_update_progress(self, percent, text):
        self.root.after(0, lambda: self.progress_var_t1.set(percent))
        self.root.after(0, lambda: self.lbl_status_t1.config(text=text))

    def chon_file(self, loai):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.csv")])
        if path:
            if loai == "sk1": self.file_saoke_path = path; self.lbl_saoke.config(text=os.path.basename(path), fg="blue")
            elif loai == "mst": self.file_master_path = path; self.lbl_master.config(text=os.path.basename(path), fg="blue")
            elif loai == "muavao": self.file_muavao_path = path; self.lbl_muavao.config(text=os.path.basename(path), fg="green")
            elif loai == "banra": self.file_banra_path = path; self.lbl_banra.config(text=os.path.basename(path), fg="green")

if __name__ == "__main__":
    root = tk.Tk()
    app = AppGomNghiepVu(root)
    root.mainloop()